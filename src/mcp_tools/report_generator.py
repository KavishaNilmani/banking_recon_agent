"""
MCP Tool 13 — Report Generator
Creates a multi-sheet Excel report.

Generic mode (output_columns in config):
  Sheet 1: Recon Results      — user-specified columns + Status + Reasoning (colour-coded)
  Sheet 2: Matched List       — clean table of MATCHED rows with output_columns only
  Sheet 3: Exception Report
  Sheet 4: Bank Orphans
  Sheet 5: Audit Log
  Sheet 6: Dashboard Summary

Legacy mode (no output_columns): uses default banking column layout.
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.mcp_tools import state

GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
GREY   = PatternFill("solid", fgColor="F2F2F2")
BLUE_H = PatternFill("solid", fgColor="1F497D")
BROWN  = PatternFill("solid", fgColor="7B3F00")
DASH   = PatternFill("solid", fgColor="D9E1F2")
TEAL_H = PatternFill("solid", fgColor="17375E")


def _hdr(ws, row, cols, fill):
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF" if fill not in (DASH,) else "000000")
        cell.alignment = Alignment(horizontal="center")


def _fill_row(ws, row, n_cols, fill):
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).fill = fill


def _auto_width(ws):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)


def _get_val_from_record(ar: dict, custom_fill: dict, col_name: str):
    """Look up a column value — checks custom_fill first, then AR record (raw and normalised)."""
    v = custom_fill.get(col_name)
    if v is not None:
        return v
    v = ar.get(col_name)
    if v is not None:
        return v
    # Try lowercase / underscore variant (canonical name)
    v = ar.get(col_name.lower())
    if v is not None:
        return v
    return ar.get(col_name.lower().replace(" ", "_"), "")


def generate_report(output_dir: str = "output") -> dict:
    """
    Write the final reconciliation Excel report.
    Call this after log_audit.
    """
    st      = state.get()
    config  = state.get_config()
    dec     = st.get("decisions", [])
    exc     = st.get("exceptions", [])
    audit   = st.get("audit", [{}])[-1]
    ar_recs = st["classified"].get("ar", [])
    b_recs  = st["classified"].get("bank", [])

    if not ar_recs:
        return {"error": "No classified source records."}
    if not dec:
        return {"error": "No decisions recorded. Call update_records first."}

    os.makedirs(output_dir, exist_ok=True)
    ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
    ptype = config.get("payment_type", "RECON").upper()
    path  = os.path.join(output_dir, f"{ptype.lower()}_recon_{ts}.xlsx")

    wb           = Workbook()
    decision_map = {d["ar_index"]: d for d in dec}

    # Determine output columns
    output_columns = config.get("output_columns", [])
    generic_mode   = bool(output_columns)

    if generic_mode:
        recon_hdrs = output_columns + ["Status", "Agent Reasoning"]
    else:
        recon_hdrs = [
            "AR Index", "Client", "Date", "Check/Ref", "AR Amount",
            "Status", "Bank", "Bank Date", "Cleared Amt", "Open Amt",
            "Remarks 1", "Remarks 2", "Agent Reasoning",
        ]

    # -----------------------------------------------------------------------
    # Sheet 1 — Recon Results
    # -----------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Recon Results"
    _hdr(ws1, 1, recon_hdrs, BLUE_H)

    row = 2
    for i, ar in enumerate(ar_recs):
        d           = decision_map.get(i, {})
        status      = d.get("status", "PENDING")
        custom_fill = d.get("custom_fill", {})

        if generic_mode:
            vals = [_get_val_from_record(ar, custom_fill, col) for col in output_columns]
            vals += [status, d.get("reasoning", "")]
        else:
            vals = [
                i,
                ar.get("client_id") or ar.get("client") or "",
                ar.get("date", ""),
                ar.get("check") or ar.get("cust_ref") or "",
                ar.get("amount"),
                status,
                d.get("bank", ""),
                d.get("bank_date", ""),
                d.get("cleared_amount", ""),
                d.get("open_amount", ""),
                d.get("remarks_1", ""),
                d.get("remarks_2", ""),
                d.get("reasoning", ""),
            ]

        for c, v in enumerate(vals, 1):
            ws1.cell(row=row, column=c, value=v).alignment = Alignment(wrap_text=False)
        fill = GREEN if status == "MATCHED" else RED if status == "UNMATCHED" else YELLOW if status == "PARTIAL" else GREY
        _fill_row(ws1, row, len(recon_hdrs), fill)
        row += 1

    _auto_width(ws1)
    ws1.freeze_panes = "A2"

    # -----------------------------------------------------------------------
    # Sheet 2 — Matched List (generic mode only — clean summary of matches)
    # -----------------------------------------------------------------------
    if generic_mode:
        ws_match        = wb.create_sheet("Matched List")
        matched_hdrs    = output_columns + ["Reasoning"]
        _hdr(ws_match, 1, matched_hdrs, TEAL_H)

        match_row = 2
        for i, ar in enumerate(ar_recs):
            d = decision_map.get(i, {})
            if d.get("status") not in ("MATCHED", "PARTIAL"):
                continue
            custom_fill = d.get("custom_fill", {})
            vals = [_get_val_from_record(ar, custom_fill, col) for col in output_columns]
            vals.append(d.get("reasoning", ""))
            for c, v in enumerate(vals, 1):
                ws_match.cell(row=match_row, column=c, value=v)
            _fill_row(ws_match, match_row, len(matched_hdrs), GREEN)
            match_row += 1

        _auto_width(ws_match)
        ws_match.freeze_panes = "A2"
        matched_sheet_name = "Matched List"
    else:
        matched_sheet_name = None

    # -----------------------------------------------------------------------
    # Sheet 3 — Exception Report
    # -----------------------------------------------------------------------
    ws_exc  = wb.create_sheet("Exceptions")
    e_hdrs  = ["Type", "AR Index", "Bank Index", "Bank Sheet", "AR Amount", "Bank Amount", "Difference", "Detail"]
    _hdr(ws_exc, 1, e_hdrs, BROWN)

    for row_i, e in enumerate(exc, 2):
        vals = [
            e.get("type"), e.get("ar_index"), e.get("bank_index"), e.get("bank_sheet"),
            e.get("ar_amount"), e.get("bank_amount"), e.get("difference"), e.get("detail"),
        ]
        for c, v in enumerate(vals, 1):
            ws_exc.cell(row=row_i, column=c, value=v)
        _fill_row(ws_exc, row_i, len(e_hdrs), RED if "MISSING" in str(e.get("type")) else YELLOW)

    _auto_width(ws_exc)

    # -----------------------------------------------------------------------
    # Sheet 4 — Bank Orphans
    # -----------------------------------------------------------------------
    ws_orph = wb.create_sheet("Bank Orphans")
    o_hdrs  = ["Bank Sheet", "Bank Index", "Date", "Amount", "Description"]
    _hdr(ws_orph, 1, o_hdrs, BROWN)

    matched_bank_idx = {d.get("bank_index") for d in dec if d.get("bank_index") is not None}
    row_o = 2
    for b_idx, b in enumerate(b_recs):
        if b_idx not in matched_bank_idx:
            vals = [
                b.get("__bank_sheet__", ""),
                b_idx,
                b.get("date") or b.get("transaction_date") or b.get("post_date") or "",
                b.get("amount") or b.get("credit_amount"),
                b.get("description") or b.get("transaction_description") or "",
            ]
            for c, v in enumerate(vals, 1):
                ws_orph.cell(row=row_o, column=c, value=v)
            _fill_row(ws_orph, row_o, len(o_hdrs), RED)
            row_o += 1

    _auto_width(ws_orph)

    # -----------------------------------------------------------------------
    # Sheet 5 — Audit Log
    # -----------------------------------------------------------------------
    ws_audit = wb.create_sheet("Audit Log")
    ws_audit.column_dimensions["A"].width = 35
    ws_audit.column_dimensions["B"].width = 28

    audit_rows = [
        ("--- RUN INFORMATION ---",              ""),
        ("Run ID",                               audit.get("run_id", "")),
        ("Timestamp",                            audit.get("timestamp", "")),
        ("Processing Time (sec)",                audit.get("processing_time_sec")),
        ("Agent Model",                          os.environ.get("ANTHROPIC_MODEL", os.environ.get("AZURE_OPENAI_DEPLOYMENT", ""))),
        ("", ""),
        ("--- INPUT ---",                        ""),
        ("Payment Type",                         audit.get("payment_type")),
        ("Input File",                           audit.get("input_file")),
        ("Source Sheet",                         audit.get("ar_sheet")),
        ("Bank Sheets",                          str(audit.get("bank_sheets", []))),
        ("", ""),
        ("--- RESULTS ---",                      ""),
        ("Total Source Records",                 audit.get("ar_total_records")),
        ("Matched",                              audit.get("matched_count")),
        ("Partial",                              audit.get("partial_count")),
        ("Unmatched",                            audit.get("unmatched_count")),
        ("Pending (no decision)",                audit.get("pending_count")),
        ("", ""),
        ("--- AMOUNTS ---",                      ""),
        ("Matched Total",                        f"${audit.get('matched_total_amount', 0):,.2f}"),
        ("Unmatched Total",                      f"${audit.get('unmatched_total_amount', 0):,.2f}"),
        ("", ""),
        ("--- EXCEPTIONS ---",                   ""),
        *[(f"  {k}", v) for k, v in audit.get("exception_summary", {}).items()],
        ("", ""),
        ("--- GP VALIDATION ---",                ""),
        ("GP Status",                            audit.get("gp_validation")),
        ("GP Difference",                        audit.get("gp_difference")),
        ("", ""),
        ("--- STATUS ---",                       ""),
        ("Reconciliation Status",                audit.get("recon_status")),
        ("Notes",                                audit.get("notes", "")),
        ("", ""),
        ("--- COLOUR LEGEND ---",                ""),
        ("GREEN",                                "MATCHED"),
        ("RED",                                  "UNMATCHED"),
        ("YELLOW",                               "PARTIAL"),
        ("GREY",                                 "PENDING"),
    ]

    for r, (label, val) in enumerate(audit_rows, 1):
        ca = ws_audit.cell(row=r, column=1, value=label)
        cb = ws_audit.cell(row=r, column=2, value=val)
        if label.startswith("---"):
            ca.font = Font(bold=True)
            ca.fill = DASH
        legend = {"GREEN": GREEN, "RED": RED, "YELLOW": YELLOW, "GREY": GREY}
        if label in legend:
            cb.fill = legend[label]

    # -----------------------------------------------------------------------
    # Sheet 6 — Dashboard Summary
    # -----------------------------------------------------------------------
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash.column_dimensions["A"].width = 30
    ws_dash.column_dimensions["B"].width = 20

    matched_cnt   = audit.get("matched_count", 0)
    unmatched_cnt = audit.get("unmatched_count", 0)
    partial_cnt   = audit.get("partial_count", 0)
    total_ar      = audit.get("ar_total_records", 1) or 1
    match_rate    = round(matched_cnt / total_ar * 100, 1)

    dash_rows = [
        ("BILLING RECONCILIATION DASHBOARD", ""),
        ("Payment Type",        ptype),
        ("Run Date",            datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Mode",                "Generic" if generic_mode else "Legacy"),
        ("", ""),
        ("Match Rate",          f"{match_rate}%"),
        ("Matched",             matched_cnt),
        ("Partial",             partial_cnt),
        ("Unmatched",           unmatched_cnt),
        ("Total Source Records", total_ar),
        ("", ""),
        ("Matched Amount",      f"${audit.get('matched_total_amount', 0):,.2f}"),
        ("Unmatched Amount",    f"${audit.get('unmatched_total_amount', 0):,.2f}"),
        ("", ""),
        ("Reconciliation Status", audit.get("recon_status", "")),
    ]

    for r, (label, val) in enumerate(dash_rows, 1):
        ca = ws_dash.cell(row=r, column=1, value=label)
        cb = ws_dash.cell(row=r, column=2, value=val)
        if r == 1:
            ca.font = Font(bold=True, size=14)
        if label == "Match Rate":
            cb.font = Font(bold=True, size=16, color="375623")
        if label == "Reconciliation Status":
            status_fill = GREEN if val == "COMPLETE" else YELLOW if val == "PARTIAL" else RED
            cb.fill = status_fill

    _auto_width(ws_dash)

    wb.save(path)
    st["output_path"] = path

    sheets = ["Recon Results"]
    if matched_sheet_name:
        sheets.append(matched_sheet_name)
    sheets += ["Exceptions", "Bank Orphans", "Audit Log", "Dashboard"]

    return {
        "output_file":   path,
        "sheets":        sheets,
        "mode":          "generic" if generic_mode else "legacy",
        "audit_summary": {
            "matched_count":          matched_cnt,
            "unmatched_count":        unmatched_cnt,
            "partial_count":          partial_cnt,
            "matched_total_amount":   audit.get("matched_total_amount", 0),
            "unmatched_total_amount": audit.get("unmatched_total_amount", 0),
            "recon_status":           audit.get("recon_status"),
        },
        "status": "Report generated successfully.",
    }
