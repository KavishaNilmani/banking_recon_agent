"""
MCP Tool 11 — Excel Write-Back
Writes custom_fill values from matched decisions back into a copy of the original
input Excel file. Never overwrites the original — always creates a new file with
an '_updated' suffix.

Requires:
  - classify_transactions has been called (source records have __original_row__)
  - update_records has been called (decisions have custom_fill dicts)
"""

import os
import shutil

from openpyxl import load_workbook

from src.mcp_tools import state


def _resolve_path(file_path: str) -> str:
    for candidate in [file_path, os.path.join("input", os.path.basename(file_path))]:
        if os.path.exists(candidate):
            return candidate
    raise FileNotFoundError(f"Input file not found: {file_path}")


def write_back_to_excel(
    source_sheet: str = "",
    output_suffix: str = "_updated",
    header_row: int = 0,
) -> dict:
    """
    Write custom_fill values from matched decisions into a copy of the original Excel file.

    Args:
        source_sheet  : sheet to update (reads ar_sheet from config if empty)
        output_suffix : appended before .xlsx in the output filename (default '_updated')
        header_row    : 0-based header row index — must match the value used in load_data (default 0)

    The tool reads fill_columns config and decisions (with custom_fill) from state.
    Only MATCHED and PARTIAL decisions with a non-empty custom_fill are written.
    """
    st     = state.get()
    config = state.get_config()

    input_file   = config.get("input_file", "")
    source_sheet = source_sheet or config.get("ar_sheet", "")
    decisions    = st.get("decisions", [])
    ar_records   = st["classified"].get("ar", [])

    if not input_file:
        return {"error": "No input_file in config. Call parse_prompt first."}
    if not source_sheet:
        return {"error": "No source_sheet specified and ar_sheet not in config."}
    if not decisions:
        return {"error": "No decisions in state. Call update_records first."}

    try:
        src_path = _resolve_path(input_file)
    except FileNotFoundError as e:
        return {"error": str(e)}

    # Build output path: insert suffix before the extension
    base, ext = os.path.splitext(src_path)
    out_path  = f"{base}{output_suffix}{ext}"

    # Copy original — never mutate it
    shutil.copy2(src_path, out_path)

    try:
        wb = load_workbook(out_path)
    except Exception as e:
        return {"error": f"Cannot open Excel file: {e}"}

    if source_sheet not in wb.sheetnames:
        return {"error": f"Sheet '{source_sheet}' not found in workbook. Available: {wb.sheetnames}"}

    ws = wb[source_sheet]

    # Build column-name → column-index map from the header row
    # openpyxl uses 1-based row indices; header_row is 0-based so Excel row = header_row + 1
    excel_header_row = header_row + 1
    col_index: dict[str, int] = {}
    for cell in ws[excel_header_row]:
        if cell.value is not None:
            col_index[str(cell.value).strip()] = cell.column

    if not col_index:
        return {"error": f"No column headers found at Excel row {excel_header_row} in sheet '{source_sheet}'."}

    filled_count  = 0
    skipped_count = 0
    errors        = []

    for decision in decisions:
        status      = decision.get("status", "")
        custom_fill = decision.get("custom_fill") or {}

        # Only write back for matched/partial rows that have fill values
        if status not in ("MATCHED", "PARTIAL") or not custom_fill:
            skipped_count += 1
            continue

        ar_idx = decision.get("ar_index")
        if ar_idx is None or ar_idx >= len(ar_records):
            errors.append(f"ar_index {ar_idx} out of range")
            continue

        ar_rec       = ar_records[ar_idx]
        original_row = ar_rec.get("__original_row__")

        if original_row is None:
            errors.append(f"ar_index {ar_idx} has no __original_row__ — classify_transactions may need updating")
            continue

        # Excel data row = header_row (0-based) + original_row (0-based) + 2
        # (+1 for 1-based Excel, +1 to skip past header row)
        excel_data_row = header_row + int(original_row) + 2

        for col_name, value in custom_fill.items():
            col_num = col_index.get(col_name)
            if col_num is None:
                errors.append(f"Column '{col_name}' not found in sheet '{source_sheet}' header row")
                continue
            ws.cell(row=excel_data_row, column=col_num, value=value)
            filled_count += 1

    try:
        wb.save(out_path)
    except Exception as e:
        return {"error": f"Failed to save workbook: {e}"}

    return {
        "status":        "Write-back complete.",
        "output_file":   out_path,
        "source_sheet":  source_sheet,
        "filled_cells":  filled_count,
        "skipped":       skipped_count,
        "errors":        errors,
        "note": (
            f"Original file preserved at '{src_path}'. "
            f"Updated copy saved to '{out_path}'."
        ),
        "next_step": "Call log_audit, then generate_report.",
    }
